
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

def new_records(df):
    #current_date = datetime.now().strftime("%Y-%m-%d")

    # Construct filename with current date
    filename = "updated.xlsx"

    with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=0)

        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        # defining column settings
        (max_row, max_col) = df.shape
        column_settings = [{"header": col} for col in df.columns]

        # Define table range & add table
        worksheet.add_table(0, 0, max_row, max_col -1,{
            "columns": column_settings,
            "name": "Table1",
    })

def records(df):
    file_path = "records.xlsx"


    if os.path.exists(file_path):            
        excel_records = pd.read_excel(file_path, sheet_name="Sheet1", parse_dates=["Date"])
        latest_date = excel_records["Date"].max()

        df['Date'] = pd.to_datetime(df['Date'])
        new_data = df[df["Date"] > pd.to_datetime(latest_date)]
        new_data = new_data.reset_index(drop=True)

        # Format 'Date' column as yyyy-mm-dd string
        new_data['Date'] = new_data['Date'].dt.strftime('%Y-%m-%d')

        print(f"New data since {latest_date}:")
        print(new_data.head())

        if not new_data.empty:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            start_row = sheet.max_row
            mode = "a"
            sheet_exists_option = {"if_sheet_exists": "overlay"}
            write_header = False
                # Use unpacking to only include 'if_sheet_exists' when needed
            with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode, **sheet_exists_option) as writer:
                new_data.to_excel(writer, sheet_name="Sheet1", index=False, startrow=start_row,header=write_header)
            print(f"Data appended to {file_path}")
        else:
            print("there is no new data to append")
    else:
        # Format 'Date column as yyyy-mm-dd string
        df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%Y-%m-%d')

        start_row = 0
        mode = "w"
        sheet_exists_option = {}
        write_header = True
        # Use unpacking to only include 'if_sheet_exists' when needed
        with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode, **sheet_exists_option) as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=start_row,header=write_header)
        print(f"Data written to {file_path}")


def email_records_export(df):

    # file_path = "email_records.xlsx"

    # with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
    #     df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=0)

    #     workbook = writer.book
    #     worksheet = writer.sheets["Sheet1"]

    #     # defining column settings
    #     (max_row, max_col) = df.shape
    #     column_settings = [{"header": col} for col in df.columns]

    #     # Define table range & add table
    #     worksheet.add_table(0, 0, max_row, max_col -1,{
    #         "columns": column_settings,
    #         "name": "Table1",
    # })

    # print(f"Data written to {file_path}")

    file_path = "email_records.xlsx"

    # Check if the file exists and load existing sheets
    if os.path.exists(file_path):
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=0)

            # Get access to the openpyxl workbook and sheet for styling
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]

            # Add table (openpyxl uses different API; xlsxwriter tables won't work here)
            from openpyxl.worksheet.table import Table, TableStyleInfo

            # Define table range
            table_ref = f"A1:{chr(65 + df.shape[1] - 1)}{df.shape[0] + 1}"
            table = Table(displayName="Table1", ref=table_ref)

            # Add style
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            worksheet.add_table(table)

    else:
        # File doesn't exist yet – create it from scratch using xlsxwriter
        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=0)

            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]

            (max_row, max_col) = df.shape
            column_settings = [{"header": col} for col in df.columns]

            worksheet.add_table(0, 0, max_row, max_col - 1, {
                "columns": column_settings,
                "name": "Table1",
            })

    print(f"Data written to {file_path}")




def export_data(df):
    new_records(df)
    records(df)
